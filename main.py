import sys


from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog
from PySide6.QtCore import Qt
from GUI import Ui_MainWindow

import numpy as np
import PR_equation as PR

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

import matplotlib.pyplot as plt


class Application(QMainWindow):

    def __init__(self):
        super(Application, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # Подключаем кнопку к методу
        self.ui.calculate_button.clicked.connect(self.on_calculate)
        self.ui.file_button.clicked.connect(self.on_file_button_clicked)
        # Подключаем изменение значения в eq_name к методу
        # self.ui.eq_name.currentIndexChanged.connect(self.on_combobox_change)

    def on_calculate(self):
        """ ОПЦИИ """

        import_file_name = self.ui.file_button.text()
        export_type = self.ui.export_button.currentText()


        """ РЕЗУЛЬТАТЫ"""
        # жидкость - пар чекбоксы
        l_toggle =  self.ui.l_check
        w_toggle =  self.ui.w_check
        ent_l_toggle = self.ui.l_ent
        ent_w_toggle = self.ui.w_ent
        dens_l_toggle = self.ui.l_dens
        dens_w_toggle = self.ui.w_dens
        z_l_toggle = self.ui.l_z
        z_w_toggle = self.ui.w_z

        # общие чекбосы
        stable_toggle = self.ui.s_stable
        ent_toggle = self.ui.s_ent
        density_toggle = self.ui.s_density
        z_toggle = self.ui.s_z
        
        all_params =   [l_toggle,
                        w_toggle,
                        ent_l_toggle,
                        ent_w_toggle,
                        dens_l_toggle,
                        dens_w_toggle,
                        z_l_toggle,
                        z_w_toggle,
                        stable_toggle,
                        ent_toggle,
                        density_toggle,
                        z_toggle]
        chosen_params = []
        
        roundValue = self.ui.roundbox.value()
        calcul_type = self.ui.calcul_type.currentText()

        

        
        
        """ ЧИТАЕМ ДАННЫЕ С EXCEL-ФАЙЛА"""
        def load_data(filepath):
            # Открытие файла
            wb = load_workbook(filename=filepath)
            ws = wb.active
            # Создание словаря для хранения данных по колонкам
            data = {}
            # Итерация по колонкам
            for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
                # Получение заголовка колонки
                header = ws.cell(row=1, column=col_idx).value
                # Проверка, что заголовок не None
                if header is not None:
                    # Фильтрация данных, исключая None и первый элемент (заголовок)
                    col_data = [cell for cell in col[1:] if cell is not None]
                    # Сохранение данных в словарь
                    data[header] = col_data
            return data
        # Отсеиваем ненужные чекбоксы
        def check_result_checkboxes():
            for each_elem in all_params:
                if each_elem.checkState().name == "Checked":
                    chosen_params.append(each_elem.text())
            chosen_params_str = ', '.join(chosen_params)
            self.ui.textBrowser.append(f"Выбранные параметры: {chosen_params_str}")
            return chosen_params_str, chosen_params
        # Режим фазовой диаграммы
        def result_envelope():  #ГРАФИК ФУНКЦИЯ
            
            # Расстояние функция
            def distance(point1, point2):
                return np.sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)
            
            # Реализация алгоритма ближайшего соседа
            def nearest_neighbor_sorte(points):
                sorted_points = [points[0]]  # Начинаем с первой точки
                points = points[1:]  # Оставшиеся точки
                while len(points) > 0:
                    last_point = sorted_points[-1]
                    # Найти ближайшую точку к последней добавленной точке
                    distances = np.array([distance(last_point, point) for point in points])
                    nearest_index = np.argmin(distances)
                    # Добавить ближайшую точку в отсортированный список и удалить ее из списка оставшихся точек
                    sorted_points.append(points[nearest_index])
                    points = np.delete(points, nearest_index, axis=0)
                return np.array(sorted_points)

            #ГРАФИК ФУНКЦИЯ
            def create_phase_plot(x,y):
                Pressure = [item for sublist in x for item in sublist]
                Temperature = [item for sublist in y for item in sublist]
                Pressure = np.array(Pressure)
                Temperature = np.array(Temperature)
                points = np.vstack((Temperature, Pressure)).T
                sorted_points = nearest_neighbor_sorte(points)
                sorted_temperature = sorted_points[:, 0]
                sorted_pressure = sorted_points[:, 1]
                plt.figure(figsize=(10, 6))
                plt.plot(sorted_temperature, sorted_pressure, marker="o", color="blue", label="Точки")
                plt.title("Фазовая диаграмма")
                plt.xlabel("Температура, К")
                plt.ylabel("Давление, МПа")
                plt.ylim(0, 60.1)
                plt.grid(True)
                plt.legend()
                plt.xlim(0, 623.15)
                plt.show()

            data = load_data(import_file_name)
            Fluid = []
            for i in range(len(data["component_name"])):
                Fluid.append(
                    [
                        data["component_name"][i],
                        data["z"][i],
                        data["mass"][i],
                        data["Pkr"][i],
                        data["Tkr"][i],
                        data["Vkr"][i],
                        data["w"][i],
                        data["cpen"][i],
                        data["T_boil"][i],
                        data["density_liq_phase"][i],
                    ]
                )
            
            # Очищаем текст в QTextBrowser
            self.ui.textBrowser.clear()
            # Добавляем текст в QTextBrowser
            self.ui.textBrowser.append(f"Выбрано уравнение состояния: Пенга-Робинсона (PR)")
            self.ui.textBrowser.append(f"Файл: {import_file_name}")
            self.ui.textBrowser.append(f"Экспорт в: {export_type}")
            self.ui.textBrowser.append(f"Режим: {calcul_type}")
            # Шаг при выводе на панель
            if temp_dots  == 1:
                self.ui.textBrowser.append(f"Температура [К] от {envelope_from_temp} до {envelope_to_temp} с шагом {(envelope_to_temp - envelope_from_temp)/(temp_dots)}")
            else:
                self.ui.textBrowser.append(f"Температура [К] от {envelope_from_temp} до {envelope_to_temp} с шагом {(envelope_to_temp - envelope_from_temp)/(temp_dots-1)}")
            # Шаг при выводе на панель
            if pres_dots == 1:
                self.ui.textBrowser.append(f"Давление [К] от {envelope_from_pressure} до {envelope_to_pressure} с шагом {(envelope_to_pressure - envelope_from_pressure)/(pres_dots)}")
            else:
                self.ui.textBrowser.append(f"Давление [МПа] от {envelope_from_pressure} до {envelope_to_pressure} с шагом {(envelope_to_pressure - envelope_from_pressure)/(pres_dots-1)}")
                
            if phase_diag_toggle.name == "Checked":
                self.ui.textBrowser.append(f"Построение фазовой диаграммы: Да")
            else:
                self.ui.textBrowser.append(f"Построение фазовой диаграммы: Нет")       
            
            if phase_diag_toggle_full.name == "Checked":
                self.ui.textBrowser.append(f"Построение полной фазовой диаграммы: Да")
            else:
                self.ui.textBrowser.append(f"Построение полной фазовой диаграммы: Нет")
            chosen_text, chosen = check_result_checkboxes()
            temp_array = np.linspace(envelope_from_temp, envelope_to_temp, temp_dots)
            # temp_array = 293.15
            pres_array = np.linspace(envelope_from_pressure, envelope_to_pressure, pres_dots)
            # pres_array = np.arange(0.1, 60.1, 0.1)
            # temp_array = np.linspace(200, 800, 100)
            # pres_array = np.linspace(0.1, 70, 75)
            

            W_array = []
            L_array = []
            Z_v_array = []
            Z_l_array = []
            x_i_array = []
            y_i_array = []
            Stable_array = []
            m_array = []
            enthalpy_array = []
            enthalpy_w_array = []
            enthalpy_l_array = []
            Cp_array = []
            Cp_w_array = []
            Cp_l_array = []
            Cv_array = []
            Cv_w_array = []
            Cv_l_array = []
            volume_array = []
            VolumeMy_y_array = []
            VolumeMy_x_array = []
            density_array = []
            density_y_array = []
            density_x_array = []

            result_pres_array = []
            result_temp_array = []

            P_for_plot = []
            T_for_plot = []
            for temp in temp_array:
                # ГРАФИК
                if phase_diag_toggle.name == "Checked":
                    x = []
                    y = []

                for pres in pres_array:
                    # Вызов расчета, распределение параметров по переменным
                    v = PR.PR_Flash(Fluid, BIPs = None)
                    (   W,
                        Z_v,
                        Z_l,
                        x_i,
                        y_i,
                        Stable,
                        m,
                        enthalpy,
                        enthalpy_w,
                        enthalpy_l,
                        Cp,
                        Cp_w,
                        Cp_l,
                        Cv,
                        Cv_w,
                        Cv_l,
                        volume,
                        VolumeMy_y,
                        VolumeMy_x,
                        density,
                        density_y,
                        density_x,
                    ) = v.vle(pres, temp)
                    if phase_diag_toggle.name == "Checked":
                        if Stable == 0:
                            if W > 0 and W < 1:
                                x.append(pres)
                                y.append(temp)                        
                    W_array.append(W)
                    L_array.append(1-W)
                    Z_v_array.append(Z_v)
                    Z_l_array.append(Z_l)
                    x_i_array.append(x_i)
                    y_i_array.append(y_i)
                    Stable_array.append(Stable)
                    m_array.append(m)
                    enthalpy_array.append(enthalpy)
                    enthalpy_w_array.append(enthalpy_w)
                    enthalpy_l_array.append(enthalpy_l)
                    Cp_array.append(Cp)
                    Cp_w_array.append(Cp_w)
                    Cp_l_array.append(Cp_l)
                    Cv_array.append(Cv)
                    Cv_w_array.append(Cv_w)
                    Cv_l_array.append(Cv_l)
                    volume_array.append(volume)
                    VolumeMy_y_array.append(VolumeMy_y)
                    VolumeMy_x_array.append(VolumeMy_x)
                    density_array.append(density)
                    density_y_array.append(density_y)
                    density_x_array.append(density_x)

                    result_pres_array.append(pres)
                    result_temp_array.append(temp)
                    steps_counter = temp_dots * pres_dots
                if phase_diag_toggle.name == "Checked":
                    if x[-1:] != None:
                        new_x = x[:1] + x[-1:]
                        new_y = y[:1] + y[-1:]
                        P_for_plot.append(new_y)
                        T_for_plot.append(new_x)

            SLOVAR = { 'Доля жидкости': L_array,
                       'Доля пара' :  W_array,
                       'Сверхсжимаемость пара' :  Z_v_array,
                       'Сверхсжимаемость жидкости' :  Z_l_array,
                       'x1' :  x_i_array,
                       'x2' :  y_i_array,
                       'Стабильность' :  Stable_array,
                       'x3' :  m_array,
                       'Энтальпия смеси' :  enthalpy_array,
                       'Энтальпия пара' :  enthalpy_w_array,
                       'Энтальпия жидкости' :  enthalpy_l_array,
                       'x4' :  Cp_array,
                       'x5' :  Cp_w_array,
                       'x6' :  Cp_l_array,
                       'x7' :  Cv_array,     
                       'x8' :  Cv_w_array,
                       'x9' :  Cv_l_array,
                       'x10' :  volume_array,
                       'x11' :  VolumeMy_y_array,
                       'x12' :  VolumeMy_x_array,
                       'Плотность смеси' :  density_array,
                       'Плотность пара' :  density_y_array,
                       'Плотность жидкости' :  density_x_array}
            data_to_save = {}
            for each_elem in chosen:
                for key, value in SLOVAR.items():
                    if each_elem == key:
                        data_to_save[each_elem] = value
            data_to_save["Temp [K]"] = result_temp_array
            data_to_save["Pressure [MPa]"] = result_pres_array
            data_to_save["Cp"] = Cp_array
            data_to_save["Cp_w"] = Cp_w_array
            data_to_save["Cp_l"] = Cp_l_array
            wb = Workbook()
            sheet1 = wb.active
            sheet2 = wb.create_sheet("Data")
            if export_type == "Excel":
                temp_check = ((envelope_to_temp - envelope_from_temp)/temp_dots) if temp_dots == 1 else ((envelope_to_temp - envelope_from_temp)/(temp_dots -1))
                pres_check = ((envelope_to_pressure - envelope_from_pressure)/pres_dots) if pres_dots == 1 else ((envelope_to_pressure - envelope_from_pressure)/(pres_dots -1))
                data_flash_excel_export = [
                ["Уравнение состояния", "Пенга-Робинсона (PR)"],
                ["Исходные данные (файл)", import_file_name],
                ["Режим", calcul_type],
                ["Температура [К] от ", envelope_from_temp],
                ["Температура [К] до ", envelope_to_temp],
                ["Шаг температуры [К]", temp_check],
                ["Давление [МПа] от ", envelope_from_pressure],
                ["Давление [МПа] до ", envelope_to_pressure],
                ["Шаг давления [МПа]", pres_check],
                ["Округлить до", roundValue]
            ]
                # Записываем 1-ый лист в Excel-файле
                for row in data_flash_excel_export:
                    sheet1.append(row)


                # Установка ширины столбцов sheet1
                for i in range(1, len(data_flash_excel_export[0]) + 1):
                    col_letter = ''
                    while i > 0:
                        i, remainder = divmod(i - 1, 26)
                        col_letter = chr(65 + remainder) + col_letter
                    sheet1.column_dimensions[col_letter].width = 40
                data_values = data_to_save.values()

                # Записываем 2-й лист в Excel-файле    
                sheet2.append(list(data_to_save.keys()))

                # Определяем максимальное количество строк
                max_length = max(len(values) for values in data_values)

                # Запись значений в Excel
                for i in range(max_length):
                    row = []
                    for values in data_values:
                        # Проверяем, есть ли элемент на текущем индексе
                        if i < len(values):
                            row.append(values[i])
                        else:
                            row.append("")  # Если элемента нет, добавляем пустую строку
                    sheet2.append(row)  # Записываем строку в Excel
                

                # Установка ширины столбцов sheet2
                for i in range(1, len(data_to_save.keys()) + 1):
                    col_letter = ''
                    while i > 0:
                        i, remainder = divmod(i - 1, 26)
                        col_letter = chr(65 + remainder) + col_letter
                    sheet2.column_dimensions[col_letter].width = 20
                data_values = data_to_save.values()

                # Запись данных в лист
            
                wb.save('ResultsEnvelopeExcel.xlsx')
            elif export_type == "Блокнот":
                # keys_string = '| '.join(data_to_save.keys())
                keys_string = '|'.join(f"{key}" for key in data_to_save.keys())
                temp_check = ((envelope_to_temp - envelope_from_temp)/temp_dots) if temp_dots == 1 else ((envelope_to_temp - envelope_from_temp)/(temp_dots -1))
                pres_check = ((envelope_to_pressure - envelope_from_pressure)/pres_dots) if pres_dots == 1 else ((envelope_to_pressure - envelope_from_pressure)/(pres_dots -1))
                info = (f'Уравнение состояния : "Пенга-Робинсона (PR)"\n'
                        f'Исходные данные (файл): {import_file_name}\n'
                        f'Режим: {calcul_type}\n'
                        f'Температура от, [К]: {envelope_from_temp}\n'
                        f'Температура до, [К]: {envelope_to_temp}\n'
                        f'Температура, точек: {temp_dots}\n'
                        f'Температура шаг, [К]: {temp_check}\n'
                        f'Давление от, [МПа]: {envelope_from_pressure}\n'
                        f'Давление до, [МПа]: {envelope_to_pressure}\n'
                        f'Давление, точек: {pres_dots}\n'
                        f'Давление шаг, [МПа]: {pres_check}\n'
                        f'\n'
                        f'{keys_string}\n')

                i = 0
                while i < steps_counter:
                    row_values = []
                    for key, value in data_to_save.items():
                        # Форматирование значений для выравнивания
                        formatted_value = f"{round(value[i], roundValue):<{len(key)}}"  
                        row_values.append(formatted_value)
                    
                    # Объединяем значения в строку с разделителем '|'
                    info += '|'.join(row_values) + '\n'
                    i += 1

                # Открытие файла для записи
                with open('ResultsEnvelopeNotepad.txt', 'w', encoding='utf-8') as file:
                    file.write(info)  # Запись строки в файл с переносом
            else:
                pass
            if phase_diag_toggle.name == "Checked":
                create_phase_plot(T_for_plot, P_for_plot)
            
            if phase_diag_toggle_full.name =="Checked":
                temp_array = np.linspace(200, 800, 100)
                pres_array = np.linspace(0.1, 70, 75)
                P_for_plot = []
                T_for_plot = []
                for temp in temp_array:
                    # ГРАФИК
                    x = []
                    y = []
                    for pres in pres_array:
                        # Вызов расчета, распределение параметров по переменным
                        v = PR.PR_Flash(Fluid, BIPs = None)
                        (   W,
                            Z_v,
                            Z_l,
                            x_i,
                            y_i,
                            Stable,
                            m,
                            enthalpy,
                            enthalpy_w,
                            enthalpy_l,
                            Cp,
                            Cp_w,
                            Cp_l,
                            Cv,
                            Cv_w,
                            Cv_l,
                            volume,
                            VolumeMy_y,
                            VolumeMy_x,
                            density,
                            density_y,
                            density_x,
                        ) = v.vle(pres, temp)
                        if phase_diag_toggle_full.name == "Checked":
                            if Stable == 0:
                                if W > 0 and W < 1:
                                    x.append(pres)
                                    y.append(temp)
                    if x[-1:] != None:
                        new_x = x[:1] + x[-1:]
                        new_y = y[:1] + y[-1:]
                        P_for_plot.append(new_y)
                        T_for_plot.append(new_x)
                create_phase_plot(T_for_plot, P_for_plot)


        # Режим PTFlash
        def result_flash():
            data = load_data(import_file_name)
            Fluid = []
            for i in range(len(data["component_name"])):
                Fluid.append(
                    [
                        data["component_name"][i],
                        data["z"][i],
                        data["mass"][i],
                        data["Pkr"][i],
                        data["Tkr"][i],
                        data["Vkr"][i],
                        data["w"][i],
                        data["cpen"][i],
                        data["T_boil"][i],
                        data["density_liq_phase"][i],
                    ]
                )
            # Очищаем текст в QTextBrowser
            self.ui.textBrowser.clear()
            # Добавляем текст в QTextBrowser
            self.ui.textBrowser.append(f"Выбрано уравнение состояния: Пенга-Робинсона (PR)")
            self.ui.textBrowser.append(f"Файл: {import_file_name}")
            self.ui.textBrowser.append(f"Экспорт в: {export_type}")
            self.ui.textBrowser.append(f"Режим: {calcul_type}")

            self.ui.textBrowser.append(f"Выбранная температура: {temp} К")
            self.ui.textBrowser.append(f"Выбранное давление: {pres} МПа")
            chosen_text, chosen = check_result_checkboxes()

            # Вызов расчета, распределение параметров по переменным
            v = PR.PR_Flash(Fluid, BIPs = None)
            (   W,
                Z_v,
                Z_l,
                x_i,
                y_i,
                Stable,
                m,
                enthalpy,
                enthalpy_w,
                enthalpy_l,
                Cp,
                Cp_w,
                Cp_l,
                Cv,
                Cv_w,
                Cv_l,
                volume,
                VolumeMy_y,
                VolumeMy_x,
                density,
                density_y,
                density_x,
            ) = v.vle(pres, temp)            
            data_to_export = []
            SLOVAR = { 'Доля жидкости': 1-W,
                       'Доля пара' :  W,
                       'Сверхсжимаемость пара' :  Z_v,
                       'Сверхсжимаемость жидкости' :  Z_l,
                       'x1' :  x_i,
                       'x2' :  y_i,
                       'Стабильность' :  Stable,
                       'x3' :  m,
                       'Энтальпия смеси' :  enthalpy,
                       'Энтальпия пара' :  enthalpy_w,
                       'Энтальпия жидкости' :  enthalpy_l,
                       'x4' :  Cp,
                       'x5' :  Cp_w,
                       'x6' :  Cp_l,
                       'x7' :  Cv,     
                       'x8' :  Cv_w,
                       'x9' :  Cv_l,
                       'x10' :  volume,
                       'x11' :  VolumeMy_y,
                       'x12' :  VolumeMy_x,
                       'Плотность смеси' :  density,
                       'Плотность пара' :  density_y,
                       'Плотность жидкости' :  density_x}
            for each_elem in chosen:
                for key, value in SLOVAR.items():
                    if each_elem == key:
                        self.ui.textBrowser.append(f"{each_elem} =  {round(value, roundValue)}")
                        data_to_export.append([each_elem, round(value, roundValue), value])
                        # Экспорт в Excel (PTFlash)
                        if export_type == "Excel":
                            data_flash_excel_export = [
                            ["Уравнение состояния", "Пенга-Робинсона (PR)"],
                            ["Исходные данные (файл)", import_file_name],
                            ["Режим", calcul_type],
                            ["Температура, К",temp],
                            ["Давление, МПа", pres],
                            ["Округлить до:", roundValue]
                        ]
                            for each_elem in data_to_export:
                                data_flash_excel_export.append(each_elem)
                            wb = Workbook()
                            ws = wb.active
                            # Запись данных в лист
                            for row in data_flash_excel_export:
                                ws.append(row)  # Записываем строку в Excel

                            # Сохранение файла
                            wb.save('ResultsFlashExcel.xlsx')
                        elif export_type == "Блокнот": 
                            info = (f'Уравнение состояния : Пенга-Робинсона (PR)\nИсходные данные (файл): {import_file_name}\nРежим: {calcul_type}\nТемпература, К: {temp}\nДавление, МПа: {pres}\nОкруглить до: {roundValue} знака после запятой\n\n')
                            for each_elem in data_to_export:
                                info = info + f'{each_elem[0]} (огругл., уточненное): {float(each_elem[1])} , {(float(each_elem[2])) } \n'
                            # Открытие файла для записи
                            with open('ResultsFlashNotepad.txt', 'w', encoding='utf-8') as file:
                                file.write(info)  # Запись строки в файл с переносом 
                            pass
                        else:
                            pass
                


        # Какие значения берем в зависимости от выбранного режима
        if calcul_type == "Phase Envelope":
            """ PHASE ENVELOPE """
            envelope_from_temp = self.ui.from_temp.value()
            envelope_from_pressure = self.ui.from_pres.value()
            envelope_to_temp = self.ui.to_temp.value()
            envelope_to_pressure = self.ui.to_pres.value()

            temp_dots = self.ui.dots_temp.value()
            pres_dots = self.ui.dots_pres.value()

            phase_diag_toggle = self.ui.diag_check.checkState()
            phase_diag_toggle_full = self.ui.diag_check_full.checkState()

            result_envelope()
        else:
            """ PTFLASH """
            temp = self.ui.at_temp.value()
            pres = self.ui.at_pres.value()
            result_flash()

    # Открытие файлового менеджера
    def open_file_dialog(self, button):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Выберите файл", "", "Все файлы (*)", options=options)
        if file_name:
            button.setText(file_name)  # Устанавливаем текст кнопки на путь к файлу

    # Хендлер на файлбутон
    def on_file_button_clicked(self):
        self.open_file_dialog(self.ui.file_button)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Application()
    window.show()
    sys.exit(app.exec())

